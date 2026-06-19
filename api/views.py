import os  
import json
import resend
from django.shortcuts import render
from django.db import connection
from django.db.models import Q
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, serializers
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth.hashers import check_password
from rest_framework import status

# Librerías para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 🚀 CAMBIO DE LIBRERÍA: Usamos Groq en lugar de Google GenAI
from groq import Groq 

# Importación de tus modelos
from .models import Administrador, Dj, Cliente, Usuario, UsuarioSistema, Venta, Pago


# ==========================================
# CONFIGURACIÓN DE IA SEGURA (CONTRA FILTRACIONES EN GITHUB)
# ==========================================
# os.environ.get buscará de forma invisible la variable 'GROQ_API_KEY' en tu entorno.
# Así nunca dejas texto plano de tus llaves en este archivo.
GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")

resend.api_key = os.environ.get("RESEND_API_KEY")

# Inicializamos el cliente de Groq (Maneja un fallback por si no se encuentra la variable)
client_groq = Groq(api_key=GROQ_API_KEY) if GROQ_API_KEY else None


# ==========================================
# 1. PERSONALIZACIÓN DEL LOGIN (JWT)
# ==========================================
class CustomTokenObtainPairSerializer(TokenObtainPairSerializer):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields['ususis_username'] = serializers.CharField(required=False)
        if 'username' in self.fields:
            self.fields['username'].required = False

    def validate(self, attrs):
        data_json = self.context['request'].data
        
        login_input = (
            data_json.get("ususis_username") or 
            attrs.get("username") or 
            data_json.get("email")
        )
        password_input = data_json.get("password")
        
        print(f"\n--- DEBUG LOGIN ---")
        print(f"1. Input recibido: {login_input}")

        user_obj = UsuarioSistema.objects.filter(
            Q(ususis_username=login_input) | Q(usu__usu_email=login_input)
        ).first()
        
        if not user_obj:
            print("2. ERROR: Usuario no encontrado.")
            raise serializers.ValidationError("Usuario no encontrado.")

        print(f"2. Usuario encontrado en DB: {user_obj.ususis_username}")

        if not check_password(password_input, user_obj.password):
            print("3. ERROR: Contraseña incorrecta.")
            raise serializers.ValidationError("Contraseña incorrecta.")

        if not user_obj.is_active:
            print("3. ERROR: Cuenta inactiva.")
            raise serializers.ValidationError("Cuenta inactiva.")

        refresh = RefreshToken.for_user(user_obj)
        user_role = user_obj.usu.usu_rol.lower() if user_obj.usu.usu_rol else 'cliente'

        print(f"4. LOGIN EXITOSO. Bienvenido {user_obj.ususis_username} (Rol: {user_role})\n")

        return {
            'refresh': str(refresh),
            'access': str(refresh.access_token),
            'username': user_obj.ususis_username,
            'role': user_role,
            'user_id': user_obj.usu.usu_id
        }

class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer


# ==========================================
# 2. VISTA DE REGISTRO
# ==========================================
class RegistroUsuarioView(APIView):
    def post(self, request):
        from .serializers import RegistroSerializer 
        serializer = RegistroSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "✅ Usuario creado con éxito"}, status=status.HTTP_201_CREATED)
        
        print(f"❌ Error en registro: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

# ==========================================
# 3. GESTIÓN DE USUARIOS PARA EL PANEL
# ==========================================
class ListaUsuariosView(APIView):
    def get(self, request):
        usuarios = Usuario.objects.all().values(
            'usu_id', 'usu_nombres', 'usu_apellidos', 'usu_email', 'usu_rol', 'usu_telefono'
        )
        return Response(list(usuarios), status=status.HTTP_200_OK)

    def patch(self, request, pk):
        try:
            usuario = Usuario.objects.get(pk=pk)
            nuevo_rol = request.data.get('usu_rol')
            if nuevo_rol:
                usuario.usu_rol = nuevo_rol
                usuario.save()
                return Response({"msg": f"Rol actualizado a {nuevo_rol}"}, status=status.HTTP_200_OK)
            return Response({"error": "Falta el campo usu_rol"}, status=status.HTTP_400_BAD_REQUEST)
        except Usuario.DoesNotExist:
            return Response({"error": "Usuario no encontrado"}, status=status.HTTP_404_NOT_FOUND)


# ==========================================
# 4. GESTIÓN DE VENTAS (CORREGIDO)
# ==========================================
class ListaVentasView(APIView):
    def get(self, request):
        try:
            ventas_queryset = Venta.objects.all()
            lista_ventas = []
            
            for v in ventas_queryset:
                lista_ventas.append({
                    'ven_id': getattr(v, 'ven_id', None) or v.pk,
                    'cli': v.cli.pk if v.cli else None,
                    'dj': v.dj.pk if v.dj else None,
                    'adm': v.adm.pk if v.adm else None,
                    'loc': v.loc.pk if v.loc else None,
                    'paq': v.paq.pk if v.paq else None,
                    'ven_fechaevento': str(v.ven_fechaevento) if v.ven_fechaevento else None,
                    'ven_duracionhoras': v.ven_duracionhoras,
                    'ven_montototal': float(v.ven_montototal) if v.ven_montototal else 0.0,
                    'ven_estado': v.ven_estado,
                })
                
            return Response(lista_ventas, status=status.HTTP_200_OK)
            
        except Exception as e:
            print(f"❌ Error interno en ListaVentasView: {str(e)}")
            with connection.cursor() as cursor:
                cursor.execute("SELECT * FROM venta")
                columnas = [col[0].lower() for col in cursor.description]
                filas = cursor.fetchall()
                
            lista_ventas = []
            for fila in filas:
                dict_fila = dict(zip(columnas, fila))
                lista_ventas.append({
                    'ven_id': dict_fila.get('ven_id') or dict_fila.get('id'),
                    'cli': dict_fila.get('cli_id'),
                    'dj': dict_fila.get('dj_id'),
                    'adm': dict_fila.get('adm_id'),
                    'loc': dict_fila.get('loc_id') or dict_fila.get('id_local'),
                    'paq': dict_fila.get('paq_id'),
                    'ven_fechaevento': str(dict_fila.get('ven_fechaevento')),
                    'ven_duracionhoras': dict_fila.get('ven_duracionhoras'),
                    'ven_montototal': float(dict_fila.get('ven_montototal') or 0),
                    'ven_estado': dict_fila.get('ven_estado'),
                })
            return Response(lista_ventas, status=status.HTTP_200_OK)


# ==========================================
# 5. GESTIÓN DE PAGOS (ESTABILIZADO SIN ERROR DE COLUMNA)
# ==========================================
class ListaPagosView(APIView):
    def get(self, request):
        try:
            pagos_queryset = Pago.objects.all()
            lista_pagos = []
            
            for p in pagos_queryset:
                metodo_desc = "No especificado"
                if hasattr(p, 'metpag') and p.metpag:
                    metodo_desc = getattr(p.metpag, 'metpag_descripcion', "No especificado")
                elif hasattr(p, 'metpag_id') and p.metpag_id:
                    metodo_desc = f"Método {p.metpag_id}"

                lista_pagos.append({
                    'pag_id': p.pk,
                    'ven': p.ven.pk if p.ven else None,
                    'metpag': p.metpag.pk if hasattr(p, 'metpag') and p.metpag else None,
                    'metpag_descripcion': metodo_desc,
                    'pag_monto': float(p.pag_monto) if p.pag_monto else 0.0,
                    'pag_referencia': p.pag_referencia,
                    'pag_estado': p.pag_estado,
                    'pag_fecha': str(p.pag_fecha) if p.pag_fecha else None
                })
                
            return Response(lista_pagos, status=status.HTTP_200_OK)
            
        except Exception as e:
            print(f"❌ Error interno en ListaPagosView: {str(e)}")
            with connection.cursor() as cursor:
                cursor.execute("SELECT * FROM pago")
                columnas = [col[0].lower() for col in cursor.description]
                filas = cursor.fetchall()
                
            lista_pagos = []
            for fila in filas:
                dict_fila = dict(zip(columnas, fila))
                lista_pagos.append({
                    'pag_id': dict_fila.get('pag_id') or dict_fila.get('id'),
                    'ven': dict_fila.get('ven_id'),
                    'metpag': dict_fila.get('metpag_id'),
                    'metpag_descripcion': f"Método {dict_fila.get('metpag_id')}",
                    'pag_monto': float(dict_fila.get('pag_monto') or 0),
                    'pag_referencia': dict_fila.get('pag_referencia'),
                    'pag_estado': dict_fila.get('pag_estado'),
                    'pag_fecha': str(dict_fila.get('pag_fecha')),
                })
            return Response(lista_pagos, status=status.HTTP_200_OK)


# ==========================================
# 6. REPORTES INTELIGENTES (IA GRATUITA CON GROQ + LLAMA 3.3)
# ==========================================
class IAVistaPreviaView(APIView):
    def post(self, request):
        if not client_groq:
            return Response(
                {"error": "La API Key de Groq no está configurada en las variables de entorno."}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        prompt_usuario = request.data.get('prompt', '')
        if not prompt_usuario:
            return Response({"error": "El prompt es requerido"}, status=status.HTTP_400_BAD_REQUEST)

        contexto_bd = """
        Eres un asistente experto en SQL para una base de datos MySQL de una agencia de DJs y eventos. Tu única tarea es escribir una consulta SQL válida basada en la petición del usuario.
        
        REGLAS DE ORO PARA LAS TABLAS (Usa estrictamente MINÚSCULAS para los nombres de las TABLAS, pero respeta las mayúsculas exactas de sus CAMPOS):
        
        1. Tabla `usuario` (En minúsculas):
           - Campos exactos: usu_id, usu_nombres, usu_apellidos, usu_email, usu_rol, usu_telefono, usu_tipodocumento, usu_numdocumento

        2. Tabla `dj` (En minúsculas):
           - Campos exactos: DJ_ID, DJ_Especialidad, DJ_Activo, Usu_ID
           - Relación: El campo `Usu_ID` se conecta con `usuario.usu_id`.

        3. Tabla `paquete` (En minúsculas):
           - Campos exactos: Paq_ID, Paq_Nombre, Paq_Precio, Paq_Descripcion, Paq_MaxInvitados, Paq_TiempoHoras

        4. Tabla `venta` (En minúsculas):
           - Campos exactos: Ven_ID, Ven_FechaEvento, Ven_DuracionHoras, Ven_MontoTotal, Ven_Estado, Ven_FechaRegistro, Ven_Direccion, Ven_Distrito, Ven_NombreLugar, Adm_ID, Cli_ID, DJ_ID, Paq_ID
           - Relaciones importantes: 
             * Cli_ID se conecta con usuario.usu_id (Es el Cliente)
             * DJ_ID se conecta con dj.DJ_ID (Es el perfil del DJ)
             * Paq_ID se conecta con paquete.Paq_ID (El paquete elegido de servicios)

        5. Tabla `pago` (En minúsculas):
           - Campos exactos: Pag_ID, Pag_Monto, Pag_Referencia, Pag_Estado, Pag_Fecha, MetPag_ID, Ven_ID
           - Relación: Ven_ID se conecta con venta.Ven_ID

        6. Tabla `metodopago` (En minúsculas):
           - Campos exactos: MetPag_ID, MetPag_Descripcion (Ej: 'Yape', 'Plin', 'Efectivo')
           - Relación: MetPag_ID se conecta con pago.MetPag_ID

        REGLAS CRÍTICAS DE BÚSQUEDA Y RETORNO:
        - Si te piden información sobre qué servicios o paquetes se vendieron más, realiza un INNER JOIN entre `venta` y `paquete` usando `venta.Paq_ID = paquete.Paq_ID`.
        - Si te piden la lista de paquetes disponibles o precios de paquetes, consulta directamente a la tabla `paquete` respetando sus campos reales (`Paq_Nombre`, `Paq_Precio`, etc.).
        - Si el usuario te pide una fecha como "24 de mayo" sin año, busca usando el mes y el día (Ej: WHERE MONTH(Ven_FechaEvento) = 5 AND DAY(Ven_FechaEvento) = 24).
        - Debes responder ÚNICAMENTE con un objeto JSON válido. No saludes, no uses marcas de bloque de código markdown como ```json o texto explicativo adicional.
        - Estructura exacta a retornar: {"sql": "SELECT ... QUERY SQL AQUÍ"}
        - Usa alias amigables en español en el SELECT (ej: SELECT Paq_Nombre AS 'Paquete'...) para que React dibuje la tabla de forma impecable.
        """

        try:
            # 🚀 MODELO ACTUALIZADO: Cambiado a llama-3.3-70b-versatile (Vigente y ultra veloz)
            completion = client_groq.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[
                    {"role": "system", "content": contexto_bd},
                    {"role": "user", "content": f"Genera la consulta SQL para: {prompt_usuario}"}
                ],
                temperature=0.1
            )

            respuesta_ia = completion.choices[0].message.content.strip()

            # Mantenemos las limpiezas de seguridad por si acaso
            if "<think>" in respuesta_ia:
                respuesta_ia = respuesta_ia.split("</think>")[-1].strip()

            if "```" in respuesta_ia:
                if "json" in respuesta_ia:
                    respuesta_ia = respuesta_ia.split("```json")[-1].split("```")[0].strip()
                else:
                    respuesta_ia = respuesta_ia.split("```")[1].split("```")[0].strip()

            datos_json = json.loads(respuesta_ia)
            query_sql = datos_json.get("sql")
            
            print(f"\n🔍 [SQL GENERADO POR LLAMA VIA GROQ]: {query_sql}\n")

            # --- EJECUTAR CONSULTA DINÁMICA ---
            with connection.cursor() as cursor:
                cursor.execute(query_sql)
                columnas = [col[0] for col in cursor.description]
                filas = cursor.fetchall()

            return Response({
                "columnas": columnas,
                "filas": filas
            }, status=status.HTTP_200_OK)

        except json.JSONDecodeError:
            print(f"❌ Falló el parseo de la respuesta de la IA: {respuesta_ia}")
            return Response({"error": "La IA no devolvió un formato de consulta estructurado válido."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except Exception as e:
            print("❌ Error en Reporte con Groq:", str(e))
            return Response({"error": "No se pudo procesar la consulta inteligente en el servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==========================================
# 7. EXPORTACIÓN PROFESIONAL A EXCEL
# ==========================================
class IADescargarExcelView(APIView):
    def post(self, request):
        titulo = request.data.get('titulo', 'Reporte Inteligente')
        columnas = request.data.get('columnas', [])
        filas = request.data.get('filas', [])

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Filtrado"

        ws.merge_cells('A1:E1')
        ws['A1'] = titulo
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color='217346')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30

        ws.append([]) 
        ws.append(columnas) 

        fill_header = PatternFill(start_color='217346', end_color='217346', fill_type='solid')
        font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        
        for cell in ws[3]:
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[3].height = 24

        font_datos = Font(name='Arial', size=10)
        for fila in filas:
            ws.append(fila)
            for cell in ws[ws.max_row]:
                cell.font = font_datos
                cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[ws.max_row].height = 20

        for col in ws.columns:
            max_len = 0
            for cell in col:
                valor_str = str(cell.value) if cell.value is not None else ""
                if len(valor_str) > max_len:
                    max_len = len(valor_str)
            
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Reporte_Inteligente.xlsx"'
        
        wb.save(response)
        return response
    
# ==========================================
# 8. VISTA DE REGISTRO CON ENVÍO DE CORREO (RESEND)
# ==========================================
class RegistroUsuarioView(APIView):
    def post(self, request):
        from .serializers import RegistroSerializer 
        serializer = RegistroSerializer(data=request.data)
        
        if serializer.is_valid():
            # Carga la API Key en secreto desde las variables de entorno de Render
            resend.api_key = os.getenv("RESEND_API_KEY")
            
            # 1. Guarda el usuario en la base de datos
            usuario_sistema = serializer.save()
            
            # 2. Extraer datos limpios para el correo
            correo_destino = request.data.get('email')
            nombre_usuario = request.data.get('nombre', 'Usuario')
            
            # 3. Intentar enviar el correo de bienvenida usando Resend
            try:
                # Simplificamos la estructura y el vocabulario para pasar el filtro estricto de Gmail
                html_content = f"""
                <div style="font-family: Arial, sans-serif; padding: 20px; max-width: 600px; margin: 0 auto; border: 1px solid #eeeeee;">
                    <h2>¡Hola, {nombre_usuario}!</h2>
                    <p>Gracias por registrarte en nuestra plataforma de Bambu Eventos.</p>
                    <p>Tu cuenta ha sido creada con éxito. Pronto podrás gestionar tus eventos y revisar la disponibilidad de nuestros servicios.</p>
                    <br>
                    <p>Saludos cordiales,</p>
                    <p><strong>El equipo de Bambu Eventos</strong></p>
                </div>
                """

                # Carga el remitente de forma segura desde las variables de Render
                remitente = os.getenv("RESEND_FROM_EMAIL", "Bambú Eventos <hola@bamboeventosperu.dpdns.org>")

                resend.Emails.send({
                    "from": remitente,
                    "to": [correo_destino],
                    "subject": "Confirmacion de registro en Bambu Eventos",  # Sin emojis ni exclamaciones para evitar bloqueos
                    "html": html_content
                })
                print(f"📩 [RESEND SUCCESS]: Correo enviado con éxito a {correo_destino}")
                
            except Exception as e:
                print(f"❌ [RESEND ERROR]: No se pudo despachar el correo. Motivo: {str(e)}")

            return Response({
                "message": "✅ Usuario creado con éxito. Revisa tu correo electrónico."
            }, status=status.HTTP_201_CREATED)
        
        print(f"❌ Error en registro: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)